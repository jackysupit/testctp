# -*- coding: utf-8 -*-
"""
    - created by jeksu -
requirements:
    pip3 install openpyxl
    pip3 install pillow
"""

import logging
_logger = logging.getLogger(__name__)

from odoo import http

from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, GradientFill, Alignment
from openpyxl.styles import Color
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl.chart import (
    BarChart,
    BarChart3D,
    PieChart,
    PieChart3D,
    ProjectedPieChart,
    Reference,
    DoughnutChart,
    Series
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.series_factory import SeriesFactory

class JekXLSX():
    default_font_color = 'black'
    default_bg_color = 'white'

    x_pie_chart_height = 17
    x_bar_chart_height = 10

    style_title = 'Title'
    style_subtitle = 'Headline 4'
    style_head1 = 'Headline 1'
    style_head2 = 'Headline 2'
    style_head3 = 'Headline 3'
    style_head4 = 'Headline 4'
    border_thin = Side(border_style="thin", color="000000")
    border_double = Side(border_style="double", color="ff0000")
    style_border_thin = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
    style_border_double = Border(top=border_double, left=border_double, right=border_double, bottom=border_double)

    color_blue_1 = 'e1e5e6'

    def __init__(self, default_font_color='black', default_bg_color='white'):
        self.default_font_color = default_font_color
        self.default_bg_color = default_bg_color

    def new_workbook(self):
        workbook = Workbook()
        self.create_styles(workbook) #set default styles
        return workbook

    def create_styles(self, workbook):
        """
        styles:
            style_title
        """

        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=20)
        bd = Side(style='thick', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        workbook.add_named_style(highlight)

    def download_workbook(self, workbook, output_filename):
        from tempfile import NamedTemporaryFile
        with NamedTemporaryFile() as tmp:
            workbook.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = http.request.make_response(stream,
                 headers=[('Content-Type', 'application/vnd.ms-excel'),
                          ('Content-Disposition', 'attachment; filename={output_filename};'.format(
                              output_filename=output_filename))],
                 )
        #############################################################
        #############################################################
        #Cara ini enggak bisa jalan di beberapa server
        #Harus pakai cara di Atas
        #- Jacky
        #############################################################
        #############################################################
        # response = http.request.make_response(None,
        #          headers=[('Content-Type', 'application/vnd.ms-excel'),
        #                   ('Content-Disposition', 'attachment; filename={output_filename};'.format(
        #                       output_filename=output_filename))],
        #          )
        # workbook.save(response.stream)
        return response

    def cell(self, ws, row, column, value, style):
        # _logger.info('#### row:{} | col:{} | value:{} '.format(row, column, value))
        cell = ws.cell(row=row, column=column, value=value)
        self.set_style(cell, style)

        x_row = cell.row
        row1 = ws.row_dimensions[x_row]  # get dimension for row 1

        if style == self.style_title:
            row1.height = 30
        elif style == self.style_head1:
            row1.height = 28
        elif style == self.style_head2:
            row1.height = 26
        elif style == self.style_head3:
            row1.height = 24
        elif style == self.style_head4:
            row1.height = 22

        return cell

    #return first top left cell
    def merge_cells(self, ws, row1, col1, row2, col2, value='', style=''):
        ws.merge_cells('{}{}:{}{}'.format(get_column_letter(col1), row1, get_column_letter(col2), row2))
        cell = self.cell(ws=ws, row=row1, column=col1, value=value, style=style)

        x_row = cell.row
        row1 = ws.row_dimensions[x_row]  # get dimension for row 1

        if style == self.style_title:
            row1.height = 30
        elif style == self.style_head1:
            row1.height = 28
        elif style == self.style_head2:
            row1.height = 26
        elif style == self.style_head3:
            row1.height = 24
        elif style == self.style_head4:
            row1.height = 22


        return cell

    style_table_header = 'str_style_table_header'
    default_style_table_header = '40 % - Accent1'
    style_table_body = 'str_style_table_body'
    style_table_body_number = 'str_style_table_body_number'
    style_table_body_decimal = 'str_style_table_body_decimal'
    number_format_number = '###,##0'
    number_format_decimal = '###,##0.##'
    aligntment_top_left = Alignment(horizontal="left", vertical="top")
    aligntment_top_center = Alignment(horizontal="center", vertical="top")
    aligntment_top_right = Alignment(horizontal="right", vertical="top")

    aligntment_center_left = Alignment(horizontal="left", vertical="center")
    aligntment_center_center = Alignment(horizontal="center", vertical="center")
    aligntment_center_right = Alignment(horizontal="right", vertical="center")

    from openpyxl.styles import Font
    fontStyle = Font(size="10")
    def set_style(self, cell, style, is_style_already_set=False):
        if style == self.style_table_header:
            cell.style = self.default_style_table_header
            cell.alignment = self.aligntment_top_center
        elif style == self.style_table_body:
            cell.border = self.style_border_thin
        elif style == self.style_table_body_number:
            cell.border = self.style_border_thin
            cell.number_format = self.number_format_number
        elif style == self.style_table_body_decimal:
            cell.border = self.style_border_thin
            cell.number_format = self.number_format_decimal
        else:
            if style and not is_style_already_set:
                cell.style = style

    def font_size(self, cell, size = None):
        if size:
            font = Font(size=size)
            cell.font = font
        else:
            return cell.font.size

    def bg_color(self, cell, color = None):
        if color:
            a_fill = PatternFill(start_color=Color(color),
                                  end_color=Color(color),
                                  fill_type='solid')
            cell.fill = a_fill
        else:
            return cell.font.size

    def get_column_letter(self, index_col):
        return get_column_letter(index_col)

    def font(self, **kw):
        return Font(**kw)

    default_chart_height = 7.5
    default_chart_width = 15

    #region chart pie
    def chart_pie(self, ws, destination_cell_coordinate,
                       label_min_row, label_max_row, label_min_col, label_max_col,
                       data_min_row, data_max_row, data_min_col, data_max_col,
                       title='',from_rows=False, chart_height = default_chart_height, chart_width = default_chart_width,
                        is_3d = False, add_chart = True
                  , **kw
                  ):

        if is_3d:
            chart = PieChart3D()
        else:
            chart = PieChart()

        # print("@@@@@@@@@@@@@@@ title = {title} @@@@@@@@@@@@@@@@@@".format(title=title))
        # print("default_height={} | width:{} ||| new: {} | {}".format(chart.height, chart.width, chart_height, chart_width))
        chart.height = chart_height  # default is 7.5
        chart.width = chart_width  # default is 15

        # print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        # print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        # print("@@@@@@@@@@@@@@@ title = {title} @@@@@@@@@@@@@@@@@@".format(title=title))
        # print("labels = min_col={label_min_col}, max_col={label_max_col}, min_row={label_min_row}, max_row={label_max_row}".format(label_min_col=label_min_col,label_max_col=label_max_col,label_min_row=label_min_row,label_max_row=label_max_row,))
        # print("data = min_col={data_min_col}, max_col={data_max_col}, min_row={data_min_row}, max_row={data_max_row}".format(data_min_col=data_min_col,data_max_col=data_max_col,data_min_row=data_min_row,data_max_row=data_max_row,))
        # print("from_rows: ", from_rows)
        # print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        # print("@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")

        labels = Reference(ws, min_row=label_min_row, max_row=label_max_row, min_col=label_min_col, max_col=label_max_col,)
        data = Reference(ws, min_row=data_min_row, max_row=data_max_row, min_col=data_min_col, max_col=data_max_col,)
        chart.add_data(data, titles_from_data=False, from_rows=from_rows)
        chart.set_categories(labels)

        chart.title = title

        if add_chart:
            ws.add_chart(chart, destination_cell_coordinate)

        return chart
    #endregion chart pie

    #region chart bar
    def chart_bar(self, ws, destination_cell_coordinate,
                       label_min_row, label_max_row, label_min_col, label_max_col,
                       data_min_row, data_max_row, data_min_col, data_max_col,
                       title='',from_rows=False, chart_height = default_chart_height, chart_width = default_chart_width,
                  is_3d = False, add_chart = True, show_legend=True
                  , **kw
                  ):

        if is_3d:
            chart = BarChart3D()
        else:
            chart = BarChart()

        chart.height = chart_height  # default is 7.5
        chart.width = chart_width  # default is 15

        labels = Reference(ws, min_row=label_min_row, max_row=label_max_row, min_col=label_min_col, max_col=label_max_col,)
        data = Reference(ws, min_row=data_min_row, max_row=data_max_row, min_col=data_min_col, max_col=data_max_col,)
        chart.add_data(data, titles_from_data=False, from_rows=from_rows)
        chart.set_categories(labels)

        chart.title = title

        if not show_legend:
            chart.legend = None

        if add_chart:
            ws.add_chart(chart, destination_cell_coordinate)

        for seri in chart.series:
            seri.label = True

        return chart
    #endregion chart bar

    #region bar gauge
    def contoh_gauge_chart(self, ws):
        data = [
            ["Donut", "Pie"],
            [25, 75],
            [50, 1],
            [25, 124],
            [100],
        ]
        # based on http://www.excel-easy.com/examples/gauge-chart.html
        for row in data:
            ws.append(row)

        # First chart is a doughnut chart
        c1 = DoughnutChart(firstSliceAng=270, holeSize=50)
        c1.title = "Code coverage"
        c1.legend = None
        ref = Reference(ws, min_col=1, min_row=2, max_row=5)
        s1 = Series(ref, title_from_data=False)
        slices = [DataPoint(idx=i) for i in range(4)]
        slices[0].graphicalProperties.solidFill = "FF3300"  # red
        slices[1].graphicalProperties.solidFill = "FCF305"  # yellow
        slices[2].graphicalProperties.solidFill = "1FB714"  # green
        slices[3].graphicalProperties.noFill = True  # invisible
        s1.data_points = slices
        c1.series = [s1]


        # Second chart is a pie chart
        c2 = PieChart(firstSliceAng=270)
        c2.legend = None
        ref = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=4)
        s2 = Series(ref, title_from_data=False)
        slices = [DataPoint(idx=i) for i in range(3)]
        slices[0].graphicalProperties.noFill = True  # invisible
        slices[1].graphicalProperties.solidFill = "000000"  # black needle
        slices[2].graphicalProperties.noFill = True  # invisible
        s2.data_points = slices
        c2.series = [s2]

        c1 += c2  # combine charts
        ws.add_chart(c1, "D1")

        # Second chart is a pie chart
        c2 = PieChart(firstSliceAng=270)
        # c2.legend = None
        ref = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=4)
        s2 = Series(ref, title_from_data=False)
        # slices = [DataPoint(idx=i) for i in range(3)]
        # slices[0].graphicalProperties.noFill = True  # invisible
        # slices[1].graphicalProperties.solidFill = "000000"  # black needle
        # slices[2].graphicalProperties.noFill = True  # invisible
        # s2.data_points = slices
        c2.series = [s2]

        ws.add_chart(c2, "D20")

    #endregion chart gauge
